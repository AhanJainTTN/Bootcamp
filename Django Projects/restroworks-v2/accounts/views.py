from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.views.generic.edit import FormView
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from customers.models import Customer
from django.db import IntegrityError, transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from accounts.forms import CustomerForm, CustomerAuthenticationForm, ExcelForm


# Lifecycle of a FormView:
# 1. User requests the form page (GET /upload-excel/).
# 2. Django initializes the FormView.
# 3. get() method runs, calling get_context_data().
# 4. The form is rendered in the template.
# 5. User submits the form (POST request) with data.
# 6. Django runs form_valid() if validation passes, or form_invalid() if it fails.
# If valid: Custom processing logic runs. Django redirects to get_success_url(), unless overridden.
# If invalid: Re-renders the form with error messages.
class BulkUploadForm(FormView):
    template_name = "bulk_upload.html"
    form_class = ExcelForm
    success_url = "home"

    def form_valid(self, form):
        excel_file = form.cleaned_data["excel_file"]
        expected_headers = (
            "username",
            "password",
            "first_name",
            "last_name",
            "email",
            "phone",
        )

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            sheet_headers = tuple(cell.value for cell in ws[1])

            if sheet_headers != expected_headers:
                form.add_error(
                    None,
                    f"Sheet headers do not match the expected format: {expected_headers}",
                )
                return self.form_invalid(form)

            curr_row = 1
            created_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                curr_row += 1
                # Skip headers
                username, password, first_name, last_name, email, phone = row

                if not all(row):
                    form.add_error(None, f"Row {curr_row}: Null values encountered.")
                    continue

                try:
                    # Check if user already exists - faster than hashing and letting the database handle it.
                    if User.objects.filter(username=username).exists():
                        raise IntegrityError("User already present in database.")

                    validate_password(password)
                    user = User(
                        username=username,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    print(email)
                    user.set_password(password)
                    with transaction.atomic():
                        user.save()
                        print(user)
                        customer = Customer.objects.create(
                            user=user, email=email, phone=phone
                        )
                        created_count += 1
                        print(customer)

                except (IntegrityError, ValidationError, Exception) as e:
                    form.add_error(None, f"Row {curr_row}: {str(e)}")

            if form.non_field_errors():
                return self.form_invalid(form)

        except Exception as e:
            form.add_error(None, f"Error processing file: {str(e)}")
            return self.form_invalid(form)

        # ensures that the default behavior of form_valid() is executed, which redirects the user to the success URL (get_success_url()).
        return super().form_valid(form)


def user_signup(request):
    form = CustomerForm()
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            if customer:
                form = CustomerForm()
                return render(request, "home.html")

    return render(request, "signup.html", {"form_data": form})


def user_login(request):
    if request.user.is_authenticated:
        return redirect("home")

    form = CustomerAuthenticationForm()
    next_url = request.GET.get("next")
    if request.method == "POST":
        form = CustomerAuthenticationForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            form = CustomerAuthenticationForm()

            return redirect(next_url) if next_url else redirect("home")

    return render(request, "login.html", {"form_data": form})


def render_home(request):
    return render(request, "home.html")


def user_logout(request):
    logout(request)
    return redirect("home")
